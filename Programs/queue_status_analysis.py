# -*- coding: utf-8 -*-
# -*- coding: us-ascii -*-

"""
功能：解析生成的队列的Excel文件，并根据需求，生成相应的近期队列统计信息等
日期：2019/12/10
作者：Bai-Yong-an
备注：需要更改的参数有 FolderName, DateRange, AnalysisDate, 手动注释来开闭相应的方法
    merge_queue_long_pending_job()方法中的图片大小也可以根据需要，手动更改
"""

import os
import glob
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
from pandas import Series, DataFrame
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as dates
import matplotlib.ticker as ticker
import seaborn as sns
import datetime
from pylab import *
mpl.rcParams['font.sans-serif'] = ['SimHei']


def append_df_to_excel(filename, df, sheet_name='Queue_Job_Extraction', startrow=None,
                       truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to an existing Excel file [filename] into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()

def merge_xlsx_files(user_feature_xlsx_files):
    """
    读取多个excel文件，并保存为一个excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "User_Feature_Collection"
    for filename in user_feature_xlsx_files:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            ws.append(values)
    return wb


class Queue_Analysis:

    def merge_queue_user_feature_collection(self):
        """
        初始化队列用户的二次特征表，并绘制队列相关性的图
        """
        # 完善队列用户的二次特征 excel
        User_Feature_Collection_Excel_Path = FolderName + '\\' + Title + '_last_' + str(DateRange) + '_days_user_feature_collection.xlsx'
        UFCE = openpyxl.load_workbook(User_Feature_Collection_Excel_Path)
        UFCE_sheet = "".join(UFCE.get_sheet_names())
        UFCE.save(User_Feature_Collection_Excel_Path)

        if UFCE_sheet == 'Queue_Job_Extraction':
            User_Feature_Collection_Excel = pd.read_excel(User_Feature_Collection_Excel_Path, header=None,
                                                          sheet_name='Queue_Job_Extraction')
            #注意，这里的相应列名应该和 user_feature 里面定义的一致
            User_Feature_Collection_Excel.columns = ['UselessID', 'User', 'Status', 'DateRange', 'Days_Recorded',
                                                     'Jobs_Recorded', 'Accuracy', 'MEAN_Memory', 'MEAN_CPU_Time',
                                                     'MEAN_Real_Time', 'MEAN_Pend', 'MEAN_Submission_Count_Recorded',
                                                     'MEAN_Submission_Count']
            User_Feature_Collection_Excel = User_Feature_Collection_Excel.drop('UselessID', axis=1)
            User_Feature_Collection_Excel.to_excel(User_Feature_Collection_Excel_Path, sheet_name='Queue_Info_Merged')
            print("完善了 " + User_Feature_Collection_Excel_Path + " 特征表···" + "\n")
        elif UFCE_sheet == 'Queue_Info_Merged':
            print("已经完成了对 " + User_Feature_Collection_Excel_Path + " 特征表的预处理···" + "\n")

        #Draw
        New_UFCE = pd.read_excel(User_Feature_Collection_Excel_Path)
        New_UFCE = New_UFCE[New_UFCE['Status'] == 'Active']

        New_UFCE = New_UFCE[['User', 'MEAN_Pend', 'MEAN_Memory','MEAN_CPU_Time', 'MEAN_Real_Time',
                             'MEAN_Submission_Count_Recorded', 'MEAN_Submission_Count','Days_Recorded','Jobs_Recorded']]
        New_UFCE.set_index(['User'], inplace=True)
        New_UFCE = New_UFCE.apply(lambda x: x.astype(float))
        # 分析队列特征值之间的相关性。
        print(New_UFCE.corr())

        Active_User = len(New_UFCE.index)
        New_UFCE_corr = New_UFCE.corr(method='spearman')

        f, ax = plt.subplots(figsize=(8, 8))
        # Generate a custom diverging colormap
        cmap = sns.diverging_palette(220, 10, as_cmap=True)
        # Generate a mask for the upper triangle
        mask = np.zeros_like(New_UFCE_corr, dtype=np.bool)
        mask[np.triu_indices_from(mask)] = False

        sns.heatmap(New_UFCE_corr, annot=True, annot_kws={'size': 12, 'weight': 'bold', 'color': 'black'}, mask=mask,
                    cmap=cmap, square=True, robust=True, center=0, linewidths=0.05, cbar_kws={"shrink": .5})

        ax.set_title('Correlation between Recent Active User\'s Job Features' + '\n'
                    + "[ DateRange: " + str(DateRange) + "   Active User Numbers: " + str(Active_User) + " ]", fontsize=20, position=(0.5, 1.05))
        # ax.invert_yaxis() #将Y轴逆序
        # ax.set_xlabel('X Label',fontsize=10)
        # ax.set_ylabel('Y Label',fontsize=10)
        ax.tick_params(axis='both', labelsize=12)  # x轴 y轴
        # ax.xaxis.tick_top() # 将Y轴刻度放置在top位置的集中方法
        # ax.set_xticklabels(ax.get_xticklabels(), rotation=0)
        # ax.set_yticklabels(ax.get_yticklabels(), rotation=0)
        # plt.show()
        f.savefig(FolderName + '\\' + Title + '_last_' + str(DateRange) + '_days_user_feature_collection.png', dpi=100, bbox_inches='tight')

    def merge_queue_job_submission_count(self):
        """
        初始化队列的作业提交数excel表，绘制总体的作业提交数图
        """
        ## 完善队列的作业提交总数 excel
        Queue_Job_Count_Excel_Path = FolderName + '\\' + Title + '_last_' + str(DateRange) + '_days_job_count.xlsx'

        QJCE = openpyxl.load_workbook(Queue_Job_Count_Excel_Path)
        QJCE_sheet = "".join(QJCE.get_sheet_names())
        QJCE.save(Queue_Job_Count_Excel_Path)

        if QJCE_sheet == 'Queue_Job_Extraction':
            Queue_Job_Count_Excel = pd.read_excel(Queue_Job_Count_Excel_Path, header=None, sheet_name='Queue_Job_Extraction')
            Queue_Job_Count_Excel.columns = ['UsefulID', 'User', 'Submit_YMD', 'Pend_Label', 'Error_Label',
                                             'Correct_Label']
            # 因为feature方法会重复调用job_count，所以这里的数据会重复，需要依据uselfulID 结合User 去重!!!——genius step
            Queue_Job_Count_Excel = Queue_Job_Count_Excel.drop_duplicates(subset=['UsefulID', 'User'], keep='first')
            Queue_Job_Count_Excel = Queue_Job_Count_Excel.drop('UsefulID', axis=1)
            Queue_Job_Count_Excel.to_excel(Queue_Job_Count_Excel_Path, sheet_name='Queue_Info_merged')
            print("完善了 " + Queue_Job_Count_Excel_Path + " 特征表···" + "\n")
        elif QJCE_sheet == 'Queue_Info_merged':
            print("已经完成了对 " + Queue_Job_Count_Excel_Path + " 特征表的预处理···" + "\n")

        #Draw
        New_QJCE = pd.read_excel(Queue_Job_Count_Excel_Path)
        New_QJCE_Job_Count = New_QJCE.groupby(by='Submit_YMD').sum()

        Active_User = len(np.unique(list(New_QJCE['User'])))
        Days_Recorded = len(New_QJCE_Job_Count.index)
        Queue_Submission_Count = New_QJCE_Job_Count.apply(lambda x: x['Error_Label'] + x['Correct_Label'], axis=1)
        #or   Queue_Submission_Count = len(New_QJCE)
        # MEAN_Submission_Count_Recorded = format(Submission_Count.sum() / Days_Recorded, '.2f')
        Queue_MEAN_Submission_Count = format(Queue_Submission_Count.sum() / (DateRange * Active_User), '.2f')

        x = New_QJCE_Job_Count.index
        # x_start = x.min() + datetime.timedelta(days=-2)
        # x_end = x.max() + datetime.timedelta(days=2)
        # 注意，这里的时间跨度是恒定的
        x_start = TimeNode + datetime.timedelta(days=-2)
        x_end = AnalysisDate + datetime.timedelta(days=2)

        y = np.array(list(New_QJCE_Job_Count['Pend_Label']))
        y1 = np.array(list(New_QJCE_Job_Count['Correct_Label']))
        y2 = np.array(list(New_QJCE_Job_Count['Error_Label']))

        plt.figure(figsize=(10, 6))
        plt.bar(x, -y, width=0.7, align='center', color='yellow', alpha=0.8)
        plt.bar(x, y1, width=0.7, align='center', color='green', bottom=y2, alpha=0.8)
        plt.bar(x, y2, width=0.7, align='center', color='red', alpha=0.8)

        plt.legend(['Pend_Count (>60s)', 'Correct_Count', 'Error_Count'], loc='upper left', fontsize=15)
        plt.title(Title + ' 队列近期作业数提交数统计' + '\n'
                  + '<' + '时间跨度: ' + str(DateRange) + '  有记录的提交天数: ' + str(Days_Recorded) + '>' + '\n'
                  + '<' + '作业提交总数：' + str(Queue_Submission_Count.sum()) + ' 活跃用户数：' + str(Active_User) + '>' + '\n'
                  + '<' + ' 队列每日人均提交作业数：' + str(Queue_MEAN_Submission_Count) + '>' + '\n', fontsize=20)
        plt.xlabel('作业提交日期', fontsize=15)
        plt.ylabel('每日任务数量统计', fontsize=15)
        plt.xticks(fontsize=15, rotation=0)
        plt.yticks(fontsize=15, rotation=0)
        plt.grid(axis='x', linestyle=':', linewidth=3)
        plt.grid(axis='y', linestyle=':', linewidth=3)

        plt.xlim(x_start, x_end)
        # 自定义刻度
        ax = plt.gca()
        ax.xaxis.set_major_locator(dates.DayLocator(interval=10))  # 主刻度为 每天
        ax.xaxis.set_major_formatter(dates.DateFormatter('\n%Y-%m-%d'))
        # plt.show()
        plt.savefig(fname=FolderName + "\\" + Title + '_last_' + str(DateRange) + '_days_job_count.png',  dpi=100,  bbox_inches='tight')

        print("完成" + Title + "队列最近" + str(DateRange) + "天的提交作业数随日期的变化分析" + "\n")

    def merge_queue_long_pending_job(self):
        """
        初始化队列的长时等待作业excel表，绘制总体的作业提交时刻统计图，分析排队高峰时间段
        """
        ## 完善队列的作业长时排队 excel
        Queue_Long_Pending_Job_Excel_Path = FolderName + '\\' + Title + '_last_' + str(DateRange) + '_days_long_pending_jobs.xlsx'
        QLPJE = openpyxl.load_workbook(Queue_Long_Pending_Job_Excel_Path)
        QLPJE_sheet = "".join(QLPJE.get_sheet_names())
        QLPJE.save(Queue_Long_Pending_Job_Excel_Path)

        if QLPJE_sheet == 'Queue_Job_Extraction':
            Queue_Long_Pending_Job_Excel = pd.read_excel(Queue_Long_Pending_Job_Excel_Path, header=None, sheet_name='Queue_Job_Extraction')
            Queue_Long_Pending_Job_Excel.columns = ['UselessID', 'User', 'Submit', 'Submit_HMS', 'Pend', 'Error', 'Correct']
            Queue_Long_Pending_Job_Excel = Queue_Long_Pending_Job_Excel.drop('UselessID', axis=1)
            # 不知道为啥会有空行，反正先删除掉！！！
            Queue_Long_Pending_Job_Excel = Queue_Long_Pending_Job_Excel.dropna()
            Queue_Long_Pending_Job_Excel.to_excel(Queue_Long_Pending_Job_Excel_Path, sheet_name='Queue_Info_merged')
            print("完善了 " + Queue_Long_Pending_Job_Excel_Path + " 特征表···" + "\n")
        elif QLPJE_sheet == 'Queue_Info_merged':
            print("已经完成了对 " + Queue_Long_Pending_Job_Excel_Path + " 特征表的预处理···" + "\n")

        New_QJPJE = pd.read_excel(Queue_Long_Pending_Job_Excel_Path)
        # Outliers is the full New_QJPJE
        Outliers = New_QJPJE[New_QJPJE['Pend'] > 60.0]
        True_Outliers = Outliers[Outliers['Error'] == 0.0]
        False_Outliers = Outliers[Outliers['Error'] > 0.0]
        # Draw
        # 如果没有值的话，绘图时会报错，这时候应该怎么做？加判断
        if len(Outliers) > 0:
            # x = Outliers['Submit_HMS']
            # 绘制一个很长很长的图——————显示具体每天的时刻
            x = Outliers['Submit']
            x_start = x.min() + datetime.timedelta(hours=-24)
            x_end = x.max() + datetime.timedelta(hours=24)

            x1 = True_Outliers['Submit']
            x2 = False_Outliers['Submit']

            y1 = True_Outliers['Pend']
            y2 = False_Outliers['Pend']

            # y2 = Outliers['Error'] + y1
            # y3 = Outliers['Correct'] + y2
            plt.figure(figsize=(200, 20))
            plt.plot(x1, y1, 'go ', x2, y2, 'r* ', markersize=25)
            plt.legend(['Correct', 'Error'], loc='upper left', fontsize=25)
            plt.title(Title + ' 队列所有长时等待作业(>60s)的提交时刻汇总分析' + '\n'
                     + '<' + '长时等待用户数: ' + str(len(np.unique(New_QJPJE['User'])))
                      + '  有记录的长时等待作业数: ' + str(len(Outliers)) + '>', fontsize=30)
            plt.xlabel('作业具体提交时刻', fontsize=20)
            plt.ylabel('长时等待作业的等待时间/s', fontsize=20)
            plt.xticks(fontsize=20, rotation=0)
            plt.yticks(fontsize=25, rotation=0)
            plt.grid(axis='x', linestyle=':', linewidth=3)
            plt.grid(axis='y', linestyle=':', linewidth=3)
            plt.xlim(x_start, x_end)

            # 标记异常数据点,等待时间超过一个小时的时候，标记一下——用户名 和 等待时间
            # df1 = Outliers[Outliers['Pend'] > 60 * 60 ]
            df1 = Outliers[Outliers['Pend'] > 60 ]
            for i in range(len(df1)):
                #Here is the 'Submit' parameter
                x = df1.iloc[i][2] # Submit
                y = df1.iloc[i][4] # Pend
                y1 = df1.iloc[i][1] # User
                # plt.text(x, y - 10, y, fontsize=20)
                plt.text(x, y - 10, "User:" + str(y1) + '\n' + str(x), fontsize=20)
            # 自定义刻度
            ax = plt.gca()
            ax.xaxis.set_major_locator(dates.HourLocator(interval=6))  # 主刻度为 每6小时
            ax.xaxis.set_major_formatter(dates.DateFormatter('\n\n %H'))
            ax.xaxis.set_minor_locator(dates.MinuteLocator(interval=180))  # 副刻度为 每60min
            ax.xaxis.set_minor_formatter(dates.DateFormatter('%M'))
            ax.yaxis.set_major_locator(ticker.MultipleLocator(10))
            ##设置y轴为对数坐标轴
            ax.set_yscale('log', nonposy='mask', subsy=[0])
            # plt.show()
        else:
            plt.figure(figsize=(100, 12))
            plt.legend(['Correct', 'Error'], loc='upper left', fontsize=25)
            plt.title(Title + ' 队列所有长时等待作业(>60s)的提交时刻汇总分析' + '\n'
                     + '<' + '长时等待用户数: ' + str(len(np.unique(New_QJPJE['User'])))
                      + '  有记录的长时等待作业数: ' + str(len(Outliers)) + '>', fontsize=30)
            plt.xlabel('作业具体提交时刻', fontsize=20)
            plt.ylabel('长时等待作业的等待时间/s', fontsize=20)
            plt.xticks(fontsize=20, rotation=0)
            plt.yticks(fontsize=25, rotation=0)
            plt.grid(axis='x', linestyle=':', linewidth=3)
            plt.grid(axis='y', linestyle=':', linewidth=3)
            plt.text(0.4, 0.5, "无长时等待作业", fontsize=30, bbox=dict(boxstyle='round,pad=0.5', fc='blue', ec='k', lw=1, alpha=0.5) )
        plt.savefig(fname=FolderName + "\\" + Title + '_last_' + str(DateRange) + '_days_long_pending_jobs.png', dpi=100, bbox_inches='tight')
        print("完成" + Title + "队列最近" + str(DateRange) + "天的长时间等待的作业统计分析" + "\n")


def main():

    qa = Queue_Analysis()

    qa.merge_queue_user_feature_collection()
    qa.merge_queue_job_submission_count()
    qa.merge_queue_long_pending_job()

if __name__ == '__main__':

    FolderName = r'D:\BYA_Project\HPC-log-analysis\HPC-DATA\JustQueue-2019'
    DateRange = 15
    AnalysisDate = datetime.date(2019, 12, 13)
    Title = os.path.basename(FolderName)
    TimeNode = AnalysisDate - datetime.timedelta(days=DateRange)

    os.chdir(FolderName)

    main()

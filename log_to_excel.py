# -*- coding: utf-8 -*-
# -*- coding: us-ascii -*-

import re
import openpyxl
from openpyxl import Workbook
import os
import pandas as pd
import datetime
from pandas import Series, DataFrame

# 解析用户的bhist日志，并根据需求，生成相应的excel数据表格

def parseString(str, lab1, lab2, allstr):
    """
    解析作业日志参数（bhist or bjobs）
    """
    # pattern = re.compile('\s+')
    # str = re.sub(pattern, '', str)
    try:
        if lab1 in str:
            new_str = str.split(lab1)[1].split(lab2)[0]
        else:
            new_str = 'None'
        allstr.append(new_str)
    except:
        pass
    return allstr

def parseTime(str, exp1, exp2, timestr):
    """
    解析并转换时间字符串
    """
    try:
        if exp1 in str:
            timestring = str.split(exp1)[0].split(exp2)[-1]
            # print(timestring)
            T = timestring[-8:]
            M = timestring[3:6]
            D = timestring.split(T)[0].split(M)[1]
            monthDic = {
                'Jan': '1',
                'Feb': '2',
                'Mar': '3',
                'Apr': '4',
                'May': '5',
                'Jun': '6',
                'Jul': '7',
                'Aug': '8',
                'Sep': '9',
                'Oct': '10',
                'Nov': '11',
                'Dec': '12'
            }
            myTime = '{}-{}-{} {}'.format('2019', monthDic[M], D, T)
        else:
            myTime = 'None'
        timestr.append(myTime)
    except:
        pass
    return timestr

####  Job Parameter related definitions
Job_ID = []
Job_Name = []
User = []
Project = []
Application = []
Status = []
Queue = []
# Start_Time = []
# Finish_Time = []
Submmit_Host = []
Requested_Resources = []
Execute_Host = []
CPU_Time = []
Memory = []
SWAP = []
# Nthread = []
# PGID = []
# PIDs = []
MAX_MEM = []
AVG_MEM = []

Parameter_List = [[Job_ID, 'Job<', '>,'],
                  [Job_Name, 'JobName<', '>,'],
                  [User, 'User<', '>'],
                  #  [Project, ',Project<', '>,'],
                  #  [Application, 'Application<', '>,'],
                  # [Status, ',Status<', '>,'],
                  #  [Queue, ',Queue<', '>,'],
                  #  [Submmit_Host, 'fromhost<', '>,'],
                  [Requested_Resources, 'RequestedResources<', '>,'],
                  [Execute_Host, 'onHost(s)<', '>,'],
                  [CPU_Time, 'CPUtimeusedis', 'seconds'],
                  # [Memory, 'MEM:', 'Gbytes;S'],
                  #  [SWAP, 'SWAP:', 'Gbytes;N'],
                  #  [MAX_MEM, 'MAXMEM:', 'Gbytes;A'],
                   [AVG_MEM, 'AVGMEM:', 'bytesSummary'],
                  ]

####  Job Status related definitions
Submmit = []
Running = []
Completed = []
Done = []

Status_List = [[Submmit, ':Submittedfromhost', '>'],
               [Running, ':Runningwith', ';'],
               [Completed, ':Completed<exit>', ';'],
               [Done, ':Donesuccessfully', ';'],
               ]

FolderName = r'D:\HPC_DATA\CST-2019'
os.chdir(FolderName)
Title = os.path.basename(os.getcwd())
SubFolderList = sorted([i for i in os.listdir(os.getcwd()) if os.path.isdir(i) and i.startswith(Title)])
print(SubFolderList)
# 循环遍历目录下的相应用户文件夹
for sub in SubFolderList:
    os.chdir(sub)
    SubTitle =  os.path.basename(os.getcwd())
    #判断是否有已经解析好的 *.xlsx 文件
    Summary = os.path.basename(sub) + ".xlsx"
    LogFile = os.path.basename(sub) + ".log"
    if os.path.exists(Summary):
        # os.remove(Summary)
        print("可能已经解析过" + sub + "的日志信息，此处跳过···")
    else:
        #读取Log的信息到内存？
        # Log_List = sorted([i for i in os.listdir(sub) if i.endswith('.log')])
        # for log in Log_List:
        #     wb = Workbook()
        #     ws = wb.active
        #     with open(log) as f:
        #         lines = f.readlines()
        #         f.close()
        if not os.path.exists(LogFile):
            print("Log文件不存在···")
        else:
            with open(LogFile) as f:
                lines = f.readlines()
                f.flush()
                f.close()

            wb = Workbook()
            ws = wb.active

            for line in lines:
                for i in range(0, len(list(Parameter_List))):
                    parseString(line, Parameter_List[i][1], Parameter_List[i][2], Parameter_List[i][0])
                for i in range(0, len(list(Status_List))):
                    parseTime(line, Status_List[i][1], Status_List[i][2], Status_List[i][0])

            for name in range(0, len(list(Parameter_List))):
                # print(Parameter_List[name][0])
                ws.append(Parameter_List[name][0])
                #remember to clear the list
                Parameter_List[name][0].clear()

            for name in range(0, len(list(Status_List))):
                # print(Status_List[name][0])
                ws.append(Status_List[name][0])
                Status_List[name][0].clear()

            wb.save(Summary)
            # lines.clear()

            # Modify the excel file to proper type
            try:
                dataset = pd.read_excel(Summary, header=None, sheet_name=None)
                ds = dataset['Sheet']
                new_ds = ds.T
                new_ds.columns = ['JobID', 'JobName', 'User', 'Requested_Resources', 'Execute_Host', 'CPU_Time',
                                  'AVG_MEM', 'Submmit', 'Running', 'Completed', 'Done']
                new_ds.to_excel(Summary, sheet_name='Sheet')
                print(sub + "的日志信息已解析完毕···" + '\n')
            except:
                print("There exists error when construct the excelfile of User " + sub)
                # wb = Workbook()
                # wb = openpyxl.load_workbook(FolderName + r'\Error_Parsing_List.xlsx')
                # ws = wb.active
                # ws.append(sub)
                # wb.save(FolderName + r'\Error_Parsing_List.xlsx')

    os.chdir(FolderName)


#记得加循环，记得加判断
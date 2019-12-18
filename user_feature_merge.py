import os
import glob
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd

def append_df_to_excel(filename, df, sheet_name='Queue_Job_Count', startrow=None,
                       truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to an existing Excel file [filename] into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # save the workbook
    writer.save()


def merge_xlsx_files(user_feature_xlsx_files):
    wb = Workbook()
    ws = wb.active
    ws.title = "User_Feature_Collection"
    for filename in user_feature_xlsx_files:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            ws.append(values)
    return wb



def merge_queue_job_count(queue_job_count_excel):
    """
    绘制总体的作业提交数图, 并将所有最近活跃的用户的图表合并显示
    """



def main():

    FolderName = r'D:\BYA_Project\HPC-log-analysis\HPC-DATA\CST-2019'
    os.chdir(FolderName)
    Title = os.path.basename(os.getcwd())
    SubFolderList = sorted([i for i in os.listdir(os.getcwd()) if os.path.isdir(i) and i.startswith(Title)])
    print(SubFolderList)

# 获取每个用户的 User_Feature Excel表的路径, 构造list
    xlsx_list = []
    for sub in SubFolderList:
        sub_xlsx = os.path.join(os.path.abspath(sub), sub + '_User_Feature.xlsx')
        xlsx_list.append(sub_xlsx)

    print("获得了所有用户的User_Feature的列表名称···" + "\n")
    print(xlsx_list)

    wb = merge_xlsx_files(xlsx_list)
    wb.save(Title + '_User_Feature_Collection.xlsx')

    Final_Excel = pd.read_excel(Title + '_User_Feature_Collection.xlsx', header=None, sheet_name=None)
    FE = Final_Excel['User_Feature_Collection']
    #注意，这里的相应列名应该和 user_feature 里面定义的一致
    FE.columns = ['User', 'Status', 'DateRange', 'Days_Recorded', 'Jobs_Recorded', 'Accuracy', 'MEAN_Memory', 'MEAN_CPU_Time',
                  'MEAN_Real_Time', 'MEAN_Pend', 'MEAN_Submmition_Count_Recorded', 'MEAN_Submmition_Count']
    FE.to_excel(Title + '_User_Feature_Collection.xlsx', sheet_name='Sheet')

    print("解析并合并生成了相应的用户作业特征表···" + "\n")


if __name__ == '__main__':
    main()


